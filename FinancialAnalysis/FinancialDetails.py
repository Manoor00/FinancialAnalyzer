# -*- coding: utf-8 -*-
"""
Created on Sun May 13 09:44:24 2018

@author: smanoor
"""

import openpyxl
import FinancialDetailsConfig
import MySqlWrapper

wb = openpyxl.load_workbook("C:\\Users\\smanoor\\Desktop\\Finance.xlsx")

#print(wb.get_sheet_names())

ws = wb['Finance']

'''d = ws.cell(row = 6, column = 2)
rng = ws['A1:B4']
t=tuple(rng)
for i in t:
    print(i[0].internal_value, i[1].internal_value)
print(d.internal_value)'''

configFile = FinancialDetailsConfig.ConfigFile()
dataTotals = {}
dataList = configFile.data
headerList = configFile.headers
formulaList = configFile.formula
headerToValue = {}
leftHeaders = []


def calculateData(key, value):
    total = 0
    rng = ws[value]
    t=tuple(rng)   
    for i in t:
        if i[1].internal_value != None:
            try:
                total = total + int(i[1].internal_value)
            except ValueError:
                    continue
    dataTotals[key] = total
    
def calculateFormulas(key):
    formulaKey = "Formula_"+key
    formula = formulaList[formulaKey]
    start = 0
    end = 0
    total = 0
    for c in formula[0]:
        if c == '+':
            dataKey = "Headers_"+formula[0][start:end].strip()
            total = total + headerToValue[dataKey]
            start = end + 1
        end = end + 1
    dataKey = "Headers_"+formula[0][start:end].strip()
    total = total + headerToValue[dataKey]
    headerToValue["Headers_"+key] = total

    
def calculateHeaders(key):
    #print(headerList[key])
    valueList = headerList[key]
    for i in valueList:
        newKey = "Headers_"+i
        try:
            #print("inside Try ", headerList[newKey])
            calculateHeaders(newKey)
            leftHeaders.append(i)
        except KeyError:
            #print("catch ", newKey)
            dataKey = "Data_"+i
            headerToValue[newKey] = dataTotals[dataKey]

for key in dataList:
   calculateData(key, dataList[key])

calculateHeaders(FinancialDetailsConfig.mainHeader)

for i in leftHeaders:
    try:
        calculateFormulas(i)
    except:
        print("Key Not found")
        continue

print(headerToValue)

mysqlWrapper = MySqlWrapper.MySqlWrapper(headerToValue)
mysqlWrapper.InsertDataIntoDb()

